import json
import os
import openpyxl
import re
import urllib.parse
import time
from datetime import datetime

excel_path = '/Users/kris/Desktop/AKI/ASTON报价软件数据库设计.xlsx'
html_path = 'index.html'
js_data_path = 'data.js'

def clean_val(v):
    if v == "#REF!" or v is None: return 0
    if isinstance(v, (int, float)): return float(v)
    try:
        match = re.search(r"(\d+\.?\d*)", str(v))
        if match: return float(match.group(1))
    except: pass
    return 0

try:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    db = {sn: [] for sn in wb.sheetnames}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.rows)
        if not rows: continue
        headers = [str(cell.value).strip() if cell.value else f"Col_{i}" for i, cell in enumerate(rows[0])]
        for row in rows[1:]:
            db[sn].append({headers[i]: row[i].value for i in range(min(len(headers), len(row)))})

    js_styles = [{"name": "0. 手动录入新款", "img": "https://via.placeholder.com/300?text=MANUAL+ENTRY", "desc": "手动输入模式", "baseLabor": 25.0, "sizes": {"CUSTOM": 0.5} }]
    for s in db.get("尺码用料矩阵", []):
        name = s.get("基础版型名称")
        if not name or str(name).strip() == "None": continue
        img = str(s.get("款式图片链接", "")).strip()
        if not img or img.lower() == "none" or img == "":
            img = f"https://via.placeholder.com/300?text={urllib.parse.quote(str(name))}"
        elif not img.startswith("http"):
            img_path = img.lstrip("/")
            if not img_path.lower().startswith("images/"): img_path = "images/" + img_path
            img = urllib.parse.quote(img_path)
        
        js_styles.append({
            "name": str(name), "img": img,
            "desc": str(s.get("描述", "") if s.get("描述") else "暂无描述"),
            "baseLabor": clean_val(s.get("基础工费", 25.0)),
            "sizes": {k: clean_val(v) for k, v in s.items() if k not in ["基础版型名称", "款式图片链接", "描述", "基础工费"] and v is not None}
        })

    js_proc = [{"name": str(p["工艺/辅料名称"]), "price": clean_val(p.get("基础单价")), "targets": str(p.get("适用版型", "通用")).replace("，", ",").split(",")} for p in db.get("工艺费率", []) if p.get("工艺/辅料名称")]
    js_fabrics = []
    for i, f in enumerate(db.get("面料库", [])):
        name = f.get("面料名称")
        if not name: continue
        js_fabrics.append({
            "id": str(f.get("ID", i+1)), "name": str(name),
            "gsm": clean_val(f.get("克重(g/㎡)", 0)), "width": clean_val(f.get("幅宽(cm)", 0)),
            "price_loose": clean_val(f.get("平方面料单价(散剪)", 0)),
            "price_roll": clean_val(f.get("平方面料单价(整卷)", 0)),
            "desc": str(f.get("描述", "")), "rollQty": clean_val(f.get("起订量（KG或米）", 20.0))
        })

    js_mods = [{"name": str(m["修正名称"]), "factor": clean_val(m.get("系数", 1.0)), "target": str(m.get("适用版型", "通用"))} for m in db.get("版型修正", []) if m.get("修正名称")]
    js_pkgs = [{"name": str(p["包装类型名称"]), "price": clean_val(p.get("单价(RMB)", 0))} for p in db.get("包装费", []) if p.get("包装类型名称")]
    
    js_sales = ["Alicia 李丽", "Caitlin 赵萌萌", "Elsa吕晓娈", "Bella白惠文", "Linn刘逍", "Allie王安冉", "Nancy王钰静", "Cecily郑晨希", "Casey裴枭君", "Cloud刘佩旺", "Tessa柴瑞蝶", "Taylor李倩", "Aria刘思彤", "Rowan张家乐", "秦宜萍", "张茹冰"]
    priority_countries = ["中国", "美国", "德国", "加拿大", "英国", "新西兰", "澳大利亚", "以色列", "法国", "挪威", "荷兰"]
    db_countries = sorted(list(set([str(c.get("国家名称", "")).strip() for c in db.get("国家库", []) if c.get("国家名称")])))
    other_countries = [c for c in db_countries if c not in priority_countries]
    final_countries = priority_countries + other_countries

    ts = int(time.time())
    db_full = { "styles": js_styles, "processes": js_proc, "modifiers": js_mods, "fabrics": js_fabrics, "sales": js_sales, "countries": final_countries, "pkgs": js_pkgs }
    with open(os.path.join('/Users/kris/Desktop/AKI/aston-quote-web/', js_data_path), 'w') as f: f.write(f"const DB = {json.dumps(db_full, ensure_ascii=False)};")

    with open(os.path.join('/Users/kris/Desktop/AKI/aston-quote-web/', html_path), 'w') as f:
        f.write(f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8"><title>ASTON 报价 V8.12 审计透明化版</title>
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <script src="https://cdn.tailwindcss.com"></script>
    <style>
        body {{ font-family: "PingFang SC", system-ui; background: #f1f5f9; color: #0f172a; }}
        .card {{ background: white; border-radius: 1.25rem; border: 1px solid #e2e8f0; }}
        .btn-active {{ background: #2563eb !important; color: white !important; border-color: #2563eb !important; }}
        .modal {{ display: none; position: fixed; inset: 0; background: rgba(15,23,42,0.8); z-index: 9999; align-items: center; justify-content: center; backdrop-filter: blur(4px); }}
        .tier-td {{ padding: 10px; background: #1e293b; color: white; font-weight: 700; border-radius: 8px; }}
        .active-tier-td {{ background: #2563eb !important; }}
        select {{ appearance: auto !important; -webkit-appearance: listbox !important; background-image: none !important; padding-right: 1rem !important; }}
    </style>
</head>
<body class="p-4 md:p-10">
    <div class="fixed top-2 right-2 z-50 bg-blue-600 text-white text-[8px] font-black px-2 py-0.5 rounded shadow-lg">V8.12 Transparency</div>

    <div id="loader" class="fixed inset-0 bg-white z-50 flex flex-col items-center justify-center">
        <div class="animate-spin rounded-full h-12 w-12 border-b-4 border-blue-600 mb-4"></div>
        <p class="text-sm font-bold text-slate-400">正在同步审计模块 V8.12...</p>
    </div>

    <div id="report-modal" class="modal">
        <div class="bg-white p-6 md:p-8 rounded-[2rem] max-w-lg w-[90%] shadow-2xl space-y-6">
            <div class="flex justify-between items-center"><h3 class="text-xl font-black">报价明细上报</h3><button onclick="closeModal()" class="text-slate-300 hover:text-slate-900 text-2xl">✕</button></div>
            <div class="bg-slate-50 p-5 rounded-2xl text-[11px] font-mono text-slate-600 border" id="report-text"></div>
            <div class="grid grid-cols-2 gap-3">
                <button onclick="copyToClipboard('summary')" class="bg-blue-600 text-white py-4 rounded-xl font-black text-xs">复制摘要</button>
                <button onclick="copyToClipboard('excel')" class="bg-emerald-600 text-white py-4 rounded-xl font-black text-xs">复制 Excel 行</button>
            </div>
        </div>
    </div>
    
    <div class="max-w-7xl mx-auto flex flex-col lg:flex-row gap-8">
        <div class="flex-1 space-y-8">
            <div class="flex flex-col md:flex-row items-start gap-8">
                <div class="flex flex-col gap-6 w-full md:w-72">
                    <div class="flex items-center gap-3"><div class="bg-blue-600 text-white p-2 rounded-xl font-black italic text-2xl px-3">ASTON</div><h1 class="text-2xl font-black text-slate-800 tracking-tighter">报价系统 V8.12</h1></div>
                    <div class="space-y-3">
                        <div class="space-y-1"><label class="text-[10px] font-black text-slate-400 uppercase">业务员</label><select id="sales-sel" onchange="calculate()" class="p-3 border-2 border-slate-200 rounded-xl text-sm font-black bg-white w-full outline-none focus:border-blue-500"></select></div>
                        <div class="space-y-1"><label class="text-[10px] font-black text-slate-400 uppercase">目的国</label><select id="country-sel" onchange="calculate()" class="p-3 border-2 border-slate-200 rounded-xl text-sm font-black bg-white w-full outline-none focus:border-blue-500"></select></div>
                        <div class="space-y-1"><label class="text-[10px] font-black text-slate-400 uppercase">项目名称</label><input type="text" id="client" oninput="calculate()" placeholder="项目/客户..." class="p-3 border-2 border-slate-200 rounded-xl text-sm font-black bg-white w-full outline-none focus:border-blue-500"></div>
                    </div>
                </div>
                <div class="flex-1 card p-8 flex flex-col md:flex-row gap-8 items-center bg-white shadow-sm relative overflow-hidden">
                    <div class="absolute top-4 right-4"><span id="quote-id" class="text-[10px] font-black bg-slate-100 text-slate-400 px-3 py-1 rounded-full">ID: --</span></div>
                    <img id="style-img" src="" class="h-64 w-64 object-contain rounded-[2rem] bg-slate-50 border shadow-inner" onerror="this.src='https://via.placeholder.com/300?text=IMAGE+ERROR'">
                    <div class="space-y-6 text-center md:text-left">
                        <p id="style-name" class="text-3xl font-black text-slate-800"></p>
                        <div class="flex flex-col gap-3 items-center md:items-start">
                            <p id="size-info" class="text-xs font-black text-blue-600 bg-blue-50 px-5 py-2 rounded-full inline-block">请选择尺码</p>
                            <p id="usage-info" class="text-xs font-black text-emerald-600 bg-emerald-50 px-5 py-2 rounded-full inline-block">核算用料: 0.000 ㎡</p>
                        </div>
                    </div>
                </div>
            </div>

            <div class="card p-8 space-y-6"><h3 class="text-xs font-black text-slate-400 uppercase tracking-widest">01. 版型选择</h3><div id="style-btns" class="grid grid-cols-2 md:grid-cols-4 gap-3"></div><div id="manual-area" class="hidden p-6 bg-blue-50 rounded-2xl grid grid-cols-2 gap-6 border-2 border-blue-100"><div><label class="text-[10px] font-black text-blue-800 uppercase">手动工费</label><input type="number" id="m-labor" value="25" oninput="calculate()" class="w-full p-3 border-2 border-blue-200 rounded-xl mt-2 font-black"></div><div><label class="text-[10px] font-black text-blue-800 uppercase">手动用料</label><input type="number" id="m-cons" value="0.5" oninput="calculate()" class="w-full p-3 border-2 border-blue-200 rounded-xl mt-2 font-black"></div></div></div>
            <div class="grid grid-cols-1 md:grid-cols-2 gap-8"><div class="card p-8 space-y-4"><h3 class="text-xs font-black text-slate-400 uppercase tracking-widest">02. 版型修正</h3><div id="mod-btns" class="grid grid-cols-2 gap-2"></div></div><div class="card p-8 space-y-4"><h3 class="text-xs font-black text-slate-400 uppercase tracking-widest">03. 尺码矩阵 (单选)</h3><div id="size-btns" class="flex flex-wrap gap-2"></div></div></div>
            <div class="card p-8 space-y-6"><h3 class="text-xs font-black text-emerald-600 uppercase tracking-widest">04. 特色工艺</h3><div id="proc-list" class="grid grid-cols-1 md:grid-cols-2 gap-3"></div></div>
            
            <div class="card p-8 space-y-6 border-orange-100 bg-orange-50/20">
                <div class="flex justify-between items-center"><h3 class="text-xs font-black text-orange-600 uppercase tracking-widest">05. 面料库 (单选)</h3><label class="flex items-center gap-2 cursor-pointer bg-orange-600 text-white px-4 py-2 rounded-full text-[10px] font-black shadow-lg shadow-orange-900/20"><input type="checkbox" id="mf-toggle" onchange="toggleManualFab()" class="w-4 h-4">手动录入新面料</label></div>
                
                <div id="mf-area" class="hidden space-y-4 p-6 bg-white rounded-2xl border-2 border-orange-200 shadow-inner">
                    <div class="grid grid-cols-3 gap-4">
                        <div><label class="text-[9px] font-black text-slate-400 uppercase">面料名称</label><input type="text" id="mf-name" value="新面料" oninput="calculate()" class="w-full p-3 border-2 border-slate-100 rounded-xl text-xs font-black mt-1"></div>
                        <div>
                            <label class="text-[9px] font-black text-slate-400 uppercase">散剪价格(RMB)</label>
                            <div class="flex items-center gap-2 mt-1">
                                <input type="radio" name="mf-price-mode" id="mf-radio-loose" checked onchange="setPriceMode('loose')" class="w-5 h-5 accent-orange-600">
                                <input type="number" id="mf-loose" value="35" oninput="calculate()" class="w-full p-3 border-2 border-slate-100 rounded-xl text-xs font-black">
                            </div>
                        </div>
                        <div>
                            <label class="text-[9px] font-black text-slate-400 uppercase">整卷价格(RMB)</label>
                            <div class="flex items-center gap-2 mt-1">
                                <input type="radio" name="mf-price-mode" id="mf-radio-roll" onchange="setPriceMode('roll')" class="w-5 h-5 accent-orange-600">
                                <input type="number" id="mf-roll" value="28" oninput="calculate()" class="w-full p-3 border-2 border-slate-100 rounded-xl text-xs font-black">
                            </div>
                        </div>
                    </div>
                    <div class="grid grid-cols-3 gap-4">
                        <div><label class="text-[9px] font-black text-slate-400 uppercase">整卷重量(KG/M)</label><input type="number" id="mf-rollqty" value="20" oninput="calculate()" class="w-full p-3 border-2 border-slate-100 rounded-xl text-xs font-black mt-1"></div>
                        <div><label class="text-[9px] font-black text-slate-400 uppercase">面料克重(G)</label><input type="number" id="mf-gsm" value="200" oninput="calculate()" class="w-full p-3 border-2 border-slate-100 rounded-xl text-xs font-black mt-1"></div>
                        <div><label class="text-[9px] font-black text-slate-400 uppercase">面料幅宽(CM)</label><input type="number" id="mf-width" value="150" oninput="calculate()" class="w-full p-3 border-2 border-slate-100 rounded-xl text-xs font-black mt-1"></div>
                    </div>
                </div>

                <div id="fab-db-area" class="space-y-4">
                    <div class="grid grid-cols-2 gap-3"><button id="mode-loose" onclick="setPriceMode('loose')" class="w-full py-4 text-xs font-black rounded-2xl bg-blue-600 text-white shadow-xl">散剪 (Loose)</button><button id="mode-roll" onclick="setPriceMode('roll')" class="w-full py-4 text-xs font-black rounded-2xl text-slate-400 bg-slate-50">整卷 (Roll)</button></div>
                    <input type="text" id="f-search" oninput="renderFabrics()" placeholder="搜索面料..." class="w-full p-4 border-2 border-slate-100 rounded-2xl text-sm font-bold bg-slate-50 outline-none">
                    <div class="overflow-hidden border-2 border-slate-100 rounded-2xl"><div class="overflow-y-auto" style="max-height: 400px;"><table class="w-full text-left border-collapse"><thead class="bg-slate-50 text-[10px] font-black uppercase text-slate-400 sticky top-0 z-10 border-b"><tr><th class="p-4 w-12"></th><th class="p-4">面料信息</th><th class="p-4">规格</th><th class="p-4 text-right" id="th-price">单价</th></tr></thead><tbody id="f-list" class="text-xs font-bold divide-y bg-white"></tbody></table></div></div>
                </div>
            </div>

            <div class="card p-8 space-y-6"><h3 class="text-xs font-black text-slate-400 uppercase tracking-widest">06. 包装方式</h3><div id="pkg-btns" class="grid grid-cols-3 gap-3"></div><div class="p-4 bg-slate-50 border rounded-xl flex justify-between items-center"><span class="text-[10px] font-black text-slate-500 uppercase">手动额外/覆盖包装费</span><input type="number" id="m-pkg" value="0" oninput="calculate()" class="w-24 p-2 border rounded-lg text-sm font-black text-right"></div></div>
        </div>

        <div class="w-full lg:w-[400px]">
            <div class="card bg-slate-900 text-white p-10 space-y-10 sticky top-10 shadow-3xl text-center border-none">
                <div><p class="text-slate-500 text-[10px] font-black uppercase tracking-[0.3em] mb-4">建议美金报价</p><div class="text-[80px] leading-none font-black tracking-tighter mb-2">$<span id="res-usd">0.00</span></div></div>
                <div class="grid grid-cols-1 gap-4 pt-4">
                    <button onclick="openReportModal()" class="w-full bg-emerald-500 hover:bg-emerald-600 text-white py-5 rounded-[1.5rem] font-black text-sm uppercase shadow-xl shadow-emerald-900/40 transition-all active:scale-95">🚀 发送记录给经理</button>
                    <div class="grid grid-cols-2 gap-3 text-left bg-slate-800 p-6 rounded-2xl border border-slate-800">
                        <div><p class="text-slate-500 text-[9px] uppercase font-black mb-1">总价 RMB</p><p id="res-rmb" class="text-xl font-black italic">¥ 0.00</p></div>
                        <div class="border-l border-slate-700 pl-4"><p class="text-emerald-500 text-[9px] uppercase font-black mb-1">MOQ</p><p id="res-moq" class="text-xl font-black italic text-emerald-400">0</p></div>
                    </div>
                </div>
                <div class="space-y-4 pt-6"><h4 class="text-[10px] font-black text-blue-500 uppercase tracking-widest text-center">现货档差报价</h4><table class="w-full border-separate border-spacing-y-2"><tbody id="tier-body"></tbody></table></div>
                
                <!-- 增强版审计面板 -->
                <div class="space-y-4 pt-6"><h4 class="text-[10px] font-black text-slate-600 uppercase tracking-widest text-center">报价及起订量审计</h4><div id="audit-box" class="bg-slate-950 p-6 rounded-2xl text-[10px] font-mono text-slate-400 text-left border border-slate-800 leading-relaxed shadow-inner"></div></div>
                
                <button onclick="window.print()" class="w-full text-slate-500 hover:text-white py-4 font-black text-[10px] uppercase border border-slate-800 rounded-2xl transition-all">打印 PDF 存档</button>
            </div>
        </div>
    </div>

    <script src="./data.js?v={ts}"></script>
    <script>
        function elSafeSet(id, val, attr='innerText') {{ const el = document.getElementById(id); if (el) el[attr] = val; }}
        function elSafeHTML(id, val) {{ const el = document.getElementById(id); if (el) el.innerHTML = val; }}

        let activeStyle=null, activeMods=[], selectedSize=null, selFabric=null, activePkg=null, priceMode='loose', currentID='';
        
        function init() {{
            try {{
                if (typeof DB === 'undefined') {{ showError("数据库加载失败，请刷新！"); return; }}
                activeStyle = DB.styles[0] || {{}}; activePkg = DB.pkgs[0] || {{ price: 0 }};
                const ss=document.getElementById('sales-sel'); if(ss) DB.sales.forEach(n=>ss.add(new Option(n,n)));
                const cs=document.getElementById('country-sel'); if(cs) DB.countries.forEach(n=>cs.add(new Option(n,n)));
                const st=document.getElementById('style-btns');
                if(st) DB.styles.forEach((s,i)=>{{
                    const b=document.createElement('button'); b.innerText=s.name;
                    b.className="p-3 text-left rounded-xl text-[10px] border-2 font-black " + (i===0?'btn-active':'bg-white text-slate-400 border-slate-100');
                    b.onclick=()=>{{
                        document.querySelectorAll('#style-btns button').forEach(x=>x.className="p-3 text-left rounded-xl text-[10px] border-2 font-black bg-white text-slate-400 border-slate-100");
                        b.className="p-3 text-left rounded-xl text-[10px] border-2 font-black btn-active";
                        activeStyle=s; 
                        const ma = document.getElementById('manual-area');
                        if(ma) ma.className=i===0?'p-6 bg-blue-50 rounded-2xl grid grid-cols-2 gap-6 border-2 border-blue-100':'hidden';
                        renderUI();
                    }};
                    st.appendChild(b);
                }});
                const sz=document.getElementById('size-btns');
                const tS = DB.styles.length > 1 ? DB.styles[1] : DB.styles[0];
                if(sz && tS && tS.sizes) Object.keys(tS.sizes).forEach(k=>{{
                    const b=document.createElement('button'); b.innerText=k;
                    b.className="size-item p-1.5 px-3 rounded-xl text-[9px] border-2 bg-white text-slate-400 font-black whitespace-nowrap overflow-hidden";
                    b.onclick=()=>{{
                        document.querySelectorAll('.size-item').forEach(x=>x.className="size-item p-1.5 px-3 rounded-xl text-[9px] border-2 bg-white text-slate-400 font-black whitespace-nowrap overflow-hidden");
                        b.className="size-item p-1.5 px-3 rounded-xl text-[9px] border-2 btn-active font-black whitespace-nowrap overflow-hidden";
                        selectedSize=k; updateSizeUI();
                    }};
                    sz.appendChild(b);
                }});
                const pb=document.getElementById('pkg-btns');
                if(pb) DB.pkgs.forEach((p,i)=>{{
                    const b=document.createElement('button'); b.innerText = p.name + " ¥" + p.price.toFixed(1);
                    b.className="p-3 rounded-xl text-[10px] border-2 font-black " + (i===0?'btn-active':'bg-white text-slate-400 border-slate-100');
                    b.onclick=()=>{{
                        document.querySelectorAll('#pkg-btns button').forEach(x=>x.className="p-3 rounded-xl text-[10px] border-2 font-black bg-white text-slate-400 border-slate-100");
                        b.className="p-3 rounded-xl text-[10px] border-2 font-black btn-active";
                        activePkg=p; calculate();
                    }};
                    pb.appendChild(b);
                }});
                renderUI(); 
                const ldr = document.getElementById('loader');
                if(ldr) ldr.style.display='none';
            }} catch (e) {{ console.error(e); showError("系统错误: " + e.message); }}
        }}

        function toggleManualFab() {{
            const isManual = document.getElementById('mf-toggle').checked;
            document.getElementById('mf-area').className = isManual ? 'space-y-4 p-6 bg-white rounded-2xl border-2 border-orange-200 shadow-inner block' : 'hidden';
            document.getElementById('fab-db-area').className = isManual ? 'opacity-30 pointer-events-none grayscale mt-4' : 'space-y-4';
            calculate();
        }}

        function setPriceMode(m) {{
            priceMode = m;
            const ml = document.getElementById('mode-loose'); if(ml) ml.className = m === 'loose' ? 'w-full py-4 text-xs font-black rounded-2xl bg-blue-600 text-white shadow-xl' : 'w-full py-4 text-xs font-black rounded-2xl text-slate-400 bg-slate-50';
            const mr = document.getElementById('mode-roll'); if(mr) mr.className = m === 'roll' ? 'w-full py-4 text-xs font-black rounded-2xl bg-blue-600 text-white shadow-xl' : 'w-full py-4 text-xs font-black rounded-2xl text-slate-400 bg-slate-50';
            const mrl = document.getElementById('mf-radio-loose'); if(mrl) mrl.checked = (m === 'loose');
            const mrr = document.getElementById('mf-radio-roll'); if(mrr) mrr.checked = (m === 'roll');
            elSafeSet('th-price', m === 'loose' ? '散剪价格 (RMB)' : '整卷价格 (RMB)');
            renderFabrics();
        }}

        function showError(msg) {{ const ldr = document.getElementById('loader'); if(ldr) ldr.innerHTML = '<div class="text-red-500 font-black p-10 bg-white rounded-3xl shadow-2xl text-center">' + msg + '</div>'; }}
        function renderUI() {{ const img = document.getElementById('style-img'); if(img) img.src = activeStyle.img; elSafeSet('style-name', activeStyle.name); renderMods(); renderProc(); renderFabrics(); calculate(); }}

        function renderMods() {{
            const g=document.getElementById('mod-btns'); if(!g) return; g.innerHTML=""; const sR=activeStyle.name.replace(/[0-9.\\s]/g,'');
            DB.modifiers.forEach(m=>{{
                const tR=m.target.replace(/[0-9.\\s]/g,'');
                if(tR==="通用"||sR.includes(tR)||tR.includes(sR)) {{
                    const b=document.createElement('button'); b.innerText=m.name; const act=activeMods.some(x=>x.name===m.name);
                    b.className="p-3 rounded-xl text-[10px] border-2 font-black " + (act?'btn-active':'bg-white text-slate-400 border-slate-100');
                    b.onclick=()=>{{ if(act) activeMods=activeMods.filter(x=>x.name!==m.name); else activeMods.push(m); renderMods(); calculate(); }};
                    g.appendChild(b);
                }}
            }});
        }}

        function renderProc() {{
            const g=document.getElementById('proc-list'); if(!g) return; g.innerHTML=""; const sR=activeStyle.name.replace(/[0-9.\\s]/g,'');
            DB.processes.forEach(p=>{{
                const targets = p.targets.map(t => t.replace(/[0-9.\\s]/g,''));
                if(targets.includes("通用") || targets.some(t => sR.includes(t) || t.includes(sR))) {{
                    const d=document.createElement('label'); d.className="p-4 bg-white border-2 border-slate-50 rounded-2xl flex justify-between items-center cursor-pointer hover:border-emerald-300 transition-all";
                    d.innerHTML='<div class="flex items-center gap-3"><input type="checkbox" onchange="calculate(true)" data-name="' + p.name + '" class="w-5 h-5 rounded-lg border-slate-300"><span class="text-xs font-black text-slate-700">' + p.name + '</span></div><span class="text-xs font-black text-emerald-600" id="price-' + p.name + '">¥' + p.price.toFixed(2) + '</span>';
                    g.appendChild(d);
                }}
            }});
        }}

        function renderFabrics() {{
            const fs = document.getElementById('f-search'); if(!fs) return; const s=fs.value.toLowerCase(); 
            const g=document.getElementById('f-list'); if(!g) return; g.innerHTML="";
            DB.fabrics.filter(f=>f.name.toLowerCase().includes(s)||f.id.includes(s)).forEach(f=>{{
                const act=selFabric?.id === f.id;
                const tr=document.createElement('tr'); tr.className="cursor-pointer hover:bg-slate-50 transition-all " + (act?'bg-orange-50/50':'');
                const pVal = priceMode === 'loose' ? f.price_loose : f.price_roll;
                tr.innerHTML='<td class="p-4"><input type="radio" name="fab-radio" ' + (act?'checked':'') + ' class="w-5 h-5 text-blue-600"></td><td class="p-4"><p class="text-[10px] text-slate-400 font-black">#' + f.id + '</p><p class="text-slate-800">' + f.name + '</p></td><td class="p-4 text-[10px] text-slate-500">' + f.gsm + 'G / ' + f.width + 'CM</td><td class="p-4 text-right text-blue-600 font-black">¥' + pVal.toFixed(2) + '</td>';
                tr.onclick=()=>{{ selFabric = f; renderFabrics(); }};
                g.appendChild(tr);
            }});
            calculate();
        }}

        function updateSizeUI() {{ elSafeSet('size-info', selectedSize ? "基准尺码: " + selectedSize : "请选择尺码矩阵"); calculate(); }}

        function calculate(updateProcColor=false) {{
            if(updateProcColor) {{ document.querySelectorAll('#proc-list input').forEach(i => {{ i.parentElement.parentElement.classList.toggle('proc-active', i.checked); }}); }}
            let qCons=0, labor=0;
            if(activeStyle.name.includes("手动")) {{ 
                qCons=parseFloat(document.getElementById('m-cons').value)||0; 
                labor=parseFloat(document.getElementById('m-labor').value)||0; 
            }} 
            else {{ if(!selectedSize) return; qCons=activeStyle.sizes[selectedSize]||0; labor=activeStyle.baseLabor; }}
            
            let fMod=1; activeMods.forEach(m=>fMod*=m.factor); let finalCons = qCons * fMod;
            elSafeSet('usage-info', "核算用料: " + finalCons.toFixed(3) + " ㎡");
            
            let isManualFab = document.getElementById('mf-toggle').checked;
            let fSum=0, mGSM=200, mRoll=20, fName='未选面料';
            
            if(isManualFab) {{
                fName = document.getElementById('mf-name').value;
                fSum = priceMode==='loose'?parseFloat(document.getElementById('mf-loose').value):parseFloat(document.getElementById('mf-roll').value);
                mGSM = parseFloat(document.getElementById('mf-gsm').value);
                mRoll = parseFloat(document.getElementById('mf-rollqty').value);
            }} else if(selFabric) {{
                fName = selFabric.name;
                fSum = priceMode==='loose'?selFabric.price_loose:selFabric.price_roll;
                mGSM = selFabric.gsm; mRoll = selFabric.rollQty;
            }}

            // MOQ 审计计算过程
            const fabricUnitUsage = (mGSM / 1000) * finalCons; 
            let theoryMoq = (mRoll / fabricUnitUsage) || 0;
            let finalMoq = Math.round(theoryMoq * 1.05);
            elSafeSet('res-moq', finalMoq);

            let extra=0; let procNames = [];
            document.querySelectorAll('#proc-list input:checked').forEach(i=>{{
                let p=DB.processes.find(x=>x.name===i.dataset.name);
                if(p) {{ 
                    let pp=p.price; 
                    if(p.name.includes("定染拉链")) {{ 
                        if(finalMoq<300) pp=2.8; else if(finalMoq<1000) pp=2.2; else if(finalMoq<3000) pp=1.6; else pp=1.2; 
                        elSafeSet('price-'+p.name, '¥'+pp.toFixed(2));
                    }} 
                    extra += pp; procNames.push(p.name);
                }}
            }});
            const pkgManual = parseFloat(document.getElementById('m-pkg').value) || 0;
            const pkgP = pkgManual > 0 ? pkgManual : (activePkg ? activePkg.price : 0);
            
            // 价格审计计算过程
            const netFabricCost = finalCons * fSum * 1.1;
            const totalRmb = (netFabricCost + labor + extra + pkgP) / 0.7;
            const finalUsd = totalRmb / 6.7;
            
            elSafeSet('res-usd', finalUsd.toFixed(2));
            elSafeSet('res-rmb', "¥ "+totalRmb.toFixed(2));

            const tb=document.getElementById('tier-body'); if(tb) {{
                tb.innerHTML="";
                const curIdx=finalMoq<100?0:finalMoq<300?1:finalMoq<500?2:3; 
                const bks=["0-99","100-299","300-499",">500"];
                let step = finalUsd < 1.5 ? 0.05 : finalUsd < 3 ? 0.1 : finalUsd < 5 ? 0.15 : finalUsd < 10 ? 0.2 : 0.25;
                bks.forEach((l,i)=>{{
                    const p = finalUsd + (curIdx - i) * step;
                    tb.innerHTML+=`<tr><td class="tier-td ${{i===curIdx?'active-tier-td':''}} text-[9px] uppercase tracking-widest">${{l}}</td><td class="text-right font-black text-2xl text-white font-mono">$${{p.toFixed(2)}}</td></tr>`;
                }});
            }}

            const auditHTML = `
                <div class="mb-4 text-emerald-500 font-black border-b border-slate-800 pb-2">■ 价格核算过程</div>
                <div class="space-y-1 mb-4">
                    <p>1. 面料净成: ${{finalCons.toFixed(3)}} ㎡ × ¥${{fSum.toFixed(2)}} × 1.1 = <span class="text-white">¥${{netFabricCost.toFixed(2)}}</span></p>
                    <p>2. 总额 RMB: (面料费${{netFabricCost.toFixed(2)}} + 工费${{labor.toFixed(1)}} + 工艺${{extra.toFixed(1)}} + 包装${{pkgP.toFixed(1)}}) / 0.7 = <span class="text-white">¥${{totalRmb.toFixed(2)}}</span></p>
                    <p>3. 美金报价: ¥${{totalRmb.toFixed(2)}} / 6.7 = <b class="text-emerald-400">$${{finalUsd.toFixed(2)}}</b></p>
                </div>
                <div class="mb-4 text-blue-500 font-black border-b border-slate-800 pb-2">■ 起订量核算过程</div>
                <div class="space-y-1">
                    <p>1. 单件克重: (${{mGSM}}G / 1000) × ${{finalCons.toFixed(3)}} ㎡ = <span class="text-white">${{fabricUnitUsage.toFixed(4)}} KG</span></p>
                    <p>2. 理论起订: ${{mRoll}} KG / ${{fabricUnitUsage.toFixed(4)}} KG = <span class="text-white">${{theoryMoq.toFixed(1)}} 件</span></p>
                    <p>3. 安全起订: ${{theoryMoq.toFixed(1)}} × 1.05 ≈ <b class="text-blue-400">${{finalMoq}} 件</b></p>
                </div>
            `;
            elSafeHTML('audit-box', auditHTML);

            const date = new Date().toISOString().slice(0,10).replace(/-/g,'');
            currentID = 'ASTON-' + date + '-' + Math.floor(Math.random()*900+100);
            elSafeSet('quote-id', 'ID: ' + currentID);
        }}

        function openReportModal() {{
            const sales = document.getElementById('sales-sel').value;
            const country = document.getElementById('country-sel').value;
            const client = document.getElementById('client').value || '未填项目';
            const usd = document.getElementById('res-usd').innerText;
            const moq = document.getElementById('res-moq').innerText;
            const fName = document.getElementById('mf-toggle').checked ? document.getElementById('mf-name').value : (selFabric ? selFabric.name : '未选面料');
            const summary = `【ASTON 报价汇报】\\n流水号: ${{currentID}}\\n业务员: ${{sales}}\\n目的国: ${{country}}\\n项目名: ${{client}}\\n版型: ${{activeStyle.name}}\\n面料: ${{fName}}\\n计价: ${{priceMode==='loose'?'散剪':'整卷'}}\\n美金单价: $${{usd}}\\n起订量: ${{moq}}件`;
            elSafeSet('report-text', summary);
            const rm = document.getElementById('report-modal'); if(rm) rm.style.display = 'flex';
        }}
        function closeModal() {{ const rm = document.getElementById('report-modal'); if(rm) rm.style.display = 'none'; }}
        function copyToClipboard(type) {{
            const sales = document.getElementById('sales-sel').value;
            const country = document.getElementById('country-sel').value;
            const client = document.getElementById('client').value || 'NA';
            const usd = document.getElementById('res-usd').innerText;
            const moq = document.getElementById('res-moq').innerText;
            const fName = document.getElementById('mf-toggle').checked ? document.getElementById('mf-name').value : (selFabric ? selFabric.name : 'NA');
            const date = new Date().toLocaleDateString();
            let t = type === 'summary' ? document.getElementById('report-text').innerText : `${{date}}\\t${{currentID}}\\t${{sales}}\\t${{country}}\\t${{client}}\\t${{activeStyle.name}}\\t${{fName}}\\t${{usd}}\\t${{moq}}`;
            navigator.clipboard.writeText(t.replace(/\\\\t/g, '\\t')).then(() => {{ alert('已成功复制！'); closeModal(); }});
        }}
        window.addEventListener('DOMContentLoaded', init);
    </script>
</body>
</html>""")
    print("✅ V8.12 Audit Transparency Update Complete.")

except Exception as e:
    print(f"❌ Error: {e}")
